import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

test_batches_dir = "/Volumes/T9/nutrition5k_dataset/imagery/test_batches"
true_metadata_file = "/Volumes/T9/nutrition5k_dataset/REMOVED_OUTLIERS/true_metadata_cafe1_cleaned.csv"
output_excel = "/Volumes/T9/nutrition5k_dataset/REMOVED_OUTLIERS/Results/MAPE_per_support.xlsx"

true_df = pd.read_csv(true_metadata_file)
true_df["dish_id"] = true_df["dish_id"].astype(str)

outlier_dish_ids = set()
for support_count in range(0, 8):
    for folder in os.listdir(test_batches_dir):
        folder_path = os.path.join(test_batches_dir, folder)
        file_path = os.path.join(folder_path, f"supporting_{support_count}.csv")
        if os.path.exists(file_path):
            df = pd.read_csv(file_path)
            df["dish_id"] = df["dish_id"].astype(str)
            joined = df.merge(true_df, on="dish_id", suffixes=("_pred", "_true"))
            joined["carbs_mape"] = np.abs((joined["carbs_pred"] - joined["carbs_true"]) / (joined["carbs_true"] + 1e-8)) * 100
            outlier_ids = joined.loc[joined["carbs_mape"] > 1000, "dish_id"]
            outlier_dish_ids.update(outlier_ids)

print("Excluded extreme outliers:")
for dish_id in sorted(outlier_dish_ids):
    print(f"- {dish_id}")
print(f"Total excluded: {len(outlier_dish_ids)}\n")

def calc_mape(pred_df, true_df):
    joined = pred_df.merge(true_df, on="dish_id", suffixes=("_pred", "_true"))
    mape = {}
    for col in ["mass", "calories", "protein", "fat", "carbs"]:
        true_col = joined[f"{col}_true"]
        pred_col = joined[f"{col}_pred"]
        mape[col] = np.mean(np.abs((pred_col - true_col) / (true_col + 1e-8))) * 100
    return mape

mape_table = []
for support_count in range(0, 8):
    all_preds = []
    for folder in os.listdir(test_batches_dir):
        folder_path = os.path.join(test_batches_dir, folder)
        if not os.path.isdir(folder_path):
            continue
        file_path = os.path.join(folder_path, f"supporting_{support_count}.csv")
        if os.path.exists(file_path):
            df = pd.read_csv(file_path)
            df["dish_id"] = df["dish_id"].astype(str)
            df = df[~df["dish_id"].isin(outlier_dish_ids)]
            all_preds.append(df)
    if not all_preds:
        continue
    pred_df = pd.concat(all_preds, ignore_index=True)
    mape = calc_mape(pred_df, true_df)

    row = {"# supporting images": support_count}
    row.update({f"{k} (%)": round(v, 2) for k, v in mape.items()})
    row["nutritional_avg (%)"] = round(np.mean([
        mape["calories"], mape["carbs"], mape["fat"]
    ]), 2)
    mape_table.append(row)

df_mape = pd.DataFrame(mape_table)
df_mape.to_excel(output_excel, index=False)

wb = load_workbook(output_excel)
ws = wb.active
fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

for col in range(2, ws.max_column + 1):
    values = [(row, ws.cell(row=row, column=col).value) for row in range(2, ws.max_row + 1)]
    min_row, _ = min(values, key=lambda x: x[1])
    ws.cell(row=min_row, column=col).fill = fill

wb.save(output_excel)
print(f"MAPE table saved with highlights to {output_excel}")
