import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

test_batches_dir = "/Volumes/T9/nutrition5k_dataset/imagery/test_batches"
true_metadata_file = "/Volumes/T9/nutrition5k_dataset/REMOVED_OUTLIERS/true_metadata_cafe1_cleaned.csv"
output_excel = "/Volumes/T9/nutrition5k_dataset/REMOVED_OUTLIERS/Results/MAE_per_support.xlsx"

true_df = pd.read_csv(true_metadata_file)
true_df["dish_id"] = true_df["dish_id"].astype(str)

outlier_dish_ids = set()
for support_count in range(0, 8):
    for folder in os.listdir(test_batches_dir):
        folder_path = os.path.join(test_batches_dir, folder)
        file_path = os.path.join(folder_path, f"supporting_{support_count}.csv")
        if os.path.exists(file_path):
            df = pd.read_csv(file_path)
            df["dish_id"] = df["dish_id"].astype(str)
            joined = df.merge(true_df, on="dish_id", suffixes=("_pred", "_true"))
            joined["carbs_mape"] = np.abs((joined["carbs_pred"] - joined["carbs_true"]) / (joined["carbs_true"] + 1e-8)) * 100
            outlier_ids = joined.loc[joined["carbs_mape"] > 1000, "dish_id"]
            outlier_dish_ids.update(outlier_ids)

print("Excluded extreme outliers:")
for dish_id in sorted(outlier_dish_ids):
    print(f"- {dish_id}")
print(f"Total excluded: {len(outlier_dish_ids)}\n")

def calc_mae_and_mape(pred_df, true_df):
    joined = pred_df.merge(true_df, on="dish_id", suffixes=("_pred", "_true"))
    mae = {}
    mape = {}
    for col in ["mass", "calories", "protein", "fat", "carbs"]:
        true_col = joined[f"{col}_true"]
        pred_col = joined[f"{col}_pred"]
        mae[col] = np.mean(np.abs(pred_col - true_col))
        mape[col] = np.mean(np.abs((pred_col - true_col) / (true_col + 1e-8))) * 100
    return mae, mape

rows = []
for support_count in range(0, 8):
    all_preds = []
    for folder in os.listdir(test_batches_dir):
        folder_path = os.path.join(test_batches_dir, folder)
        if not os.path.isdir(folder_path):
            continue
        file_path = os.path.join(folder_path, f"supporting_{support_count}.csv")
        if os.path.exists(file_path):
            df = pd.read_csv(file_path)
            df["dish_id"] = df["dish_id"].astype(str)
            df = df[~df["dish_id"].isin(outlier_dish_ids)]
            all_preds.append(df)
    if not all_preds:
        continue
    pred_df = pd.concat(all_preds, ignore_index=True)
    mae, mape = calc_mae_and_mape(pred_df, true_df)

    row = {"# supporting images": support_count}
    for key in ["calories", "mass", "fat", "carbs", "protein"]:
        formatted = f"{round(mae[key], 2)} / {round(mape[key], 1)}%"
        row[f"{key.title()} MAE"] = formatted

    macro_mae_percent = np.mean([mape["carbs"], mape["protein"], mape["fat"]])
    row["Macronutrient MAE"] = f"{round(macro_mae_percent, 1)}%"

    rows.append(row)

df = pd.DataFrame(rows)
df.to_excel(output_excel, index=False)

wb = load_workbook(output_excel)
ws = wb.active
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

for col in range(2, ws.max_column + 1):
    values = []
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=col).value
        try:
            if "/" in str(cell_value):
                val = float(str(cell_value).split("/")[0].strip())
            elif "%" in str(cell_value):
                val = float(str(cell_value).replace("%", "").strip())
            else:
                val = float(cell_value)
            values.append((row, val))
        except:
            continue

    if values:
        min_row, _ = min(values, key=lambda x: x[1])
        ws.cell(row=min_row, column=col).fill = green_fill

wb.save(output_excel)
print(f"Table saved with column-wise highlights to {output_excel}")
